import json
import os
import joblib
import openpyxl
import pandas as pd
import requests
from datetime import datetime, timedelta
from pathlib import Path
from calendar import monthrange, month_name
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.conf import settings
from django.db.models import Count, Q
from django.core.paginator import Paginator
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    UserProfile, Organization, Event, Accident, Plan, ReportFile, Invitation
)
from .forms import RegistrationForm, LoginForm, EventForm

from sklearn.cluster import KMeans

SESSION_KEY = "events_temp"

# Загружаем границы районов один раз глобально
DISTRICTS_GEOJSON_PATH = os.path.join(settings.BASE_DIR, 'static/data/ekb-districts.json')
DISTRICTS_DATA = None

def load_districts_data():
    global DISTRICTS_DATA
    if DISTRICTS_DATA is None:
        try:
            with open(DISTRICTS_GEOJSON_PATH, 'r', encoding='utf-8') as f:
                DISTRICTS_DATA = json.load(f)
            print(f"Загружены границы районов из {DISTRICTS_GEOJSON_PATH}")
        except Exception as e:
            print(f"Ошибка загрузки границ районов: {e}")
            DISTRICTS_DATA = []
    return DISTRICTS_DATA

def get_district_by_coords(lng, lat):
    """Определяет район по координатам (долгота, широта)"""
    load_districts_data()
    
    if not DISTRICTS_DATA:
        return ''
    if lng is None or lat is None:
        return ''
    
    def point_in_polygon(point, polygon_coords):
        lng, lat = point
        inside = False
        for i in range(len(polygon_coords)):
            x1, y1 = polygon_coords[i]
            x2, y2 = polygon_coords[(i + 1) % len(polygon_coords)]
            if ((y1 > lat) != (y2 > lat)) and (lng < (x2 - x1) * (lat - y1) / (y2 - y1) + x1):
                inside = not inside
        return inside
    
    # Важно: point = (долгота, широта)
    point = (lng, lat)
    
    for feature in DISTRICTS_DATA['features']:
        name = feature['properties']['name']
        geometry = feature['geometry']
        
        if geometry['type'] == 'Polygon':
            # В GeoJSON координаты хранятся как [долгота, широта]
            poly_coords = [[c[0], c[1]] for c in geometry['coordinates'][0]]
            if point_in_polygon(point, poly_coords):
                print(f"Попадание в район (полигон): {name} для точки {lng}, {lat}")
                return name
        elif geometry['type'] == 'MultiPolygon':
            for polygon in geometry['coordinates']:
                poly_coords = [[c[0], c[1]] for c in polygon[0]]
                if point_in_polygon(point, poly_coords):
                    print(f"Попадание в район (MultiPolygon): {name} для точки {lng}, {lat}")
                    return name
    
    print(f"Точка {lng}, {lat} не попала ни в один полигон")
    return ''

def get_organization_type_from_okved(okved):
    if not okved:
        return "other"

    okved_str = str(okved)

    if okved_str.startswith("85"):
        if "85.11" in okved_str:
            return "kindergarten"
        return "school"
    if okved_str.startswith("93") or okved_str.startswith("94"):
        return "sport"
    if (
        okved_str.startswith("49")
        or okved_str.startswith("50")
        or okved_str.startswith("51")
        or okved_str.startswith("52")
        or okved_str.startswith("53")
    ):
        return "transport"

    return "other"


@csrf_exempt
def search_organization_by_inn(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            inn = data.get("inn", "").strip()

            if not inn:
                return JsonResponse({"error": "Введите ИНН"}, status=400)

            if len(inn) not in [10, 12]:
                return JsonResponse(
                    {"error": "ИНН должен содержать 10 или 12 цифр"}, status=400
                )

            existing_org = Organization.objects.filter(inn=inn).first()
            if existing_org:
                return JsonResponse(
                    {
                        "found": True,
                        "name": existing_org.name,
                        "inn": existing_org.inn,
                        "ogrn": existing_org.ogrn or "",
                        "address": existing_org.address or "",
                        "organization_id": existing_org.id,
                        "from_db": True,
                    }
                )

            url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/findById/party"
            headers = {
                "Authorization": f"Token {settings.DADATA_API_KEY}",
                "Content-Type": "application/json",
            }
            payload = {"query": inn, "branch_type": "MAIN"}

            response = requests.post(url, json=payload, headers=headers, timeout=10)
            result = response.json()

            if result.get("suggestions"):
                org = result["suggestions"][0]["data"]

                new_org = Organization.objects.create(
                    name=org.get("name", {}).get("full", ""),
                    inn=org.get("inn", ""),
                    ogrn=org.get("ogrn", ""),
                    kpp=org.get("kpp", ""),
                    address=org.get("address", {}).get("value", ""),
                    is_active=True,
                )

                # Геокодируем адрес организации
                lat, lng = geocode_address(new_org.address)
                if lat and lng:
                    new_org.latitude = lat
                    new_org.longitude = lng
                    new_org.save()

                return JsonResponse(
                    {
                        "found": True,
                        "name": new_org.name,
                        "inn": new_org.inn,
                        "ogrn": new_org.ogrn or "",
                        "address": new_org.address or "",
                        "organization_id": new_org.id,
                        "from_db": False,
                    }
                )
            else:
                return JsonResponse({"found": False, "error": "Организация не найдена"})

        except requests.exceptions.Timeout:
            return JsonResponse({"error": "Превышено время ожидания"}, status=500)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Метод не разрешен"}, status=405)


def search_organizations(request):
    query = request.GET.get("q", "")
    if len(query) >= 2:
        orgs = Organization.objects.filter(
            name__icontains=query, is_active=True
        ).values("id", "name", "inn", "address")[:10]
        return JsonResponse(list(orgs), safe=False)
    return JsonResponse([], safe=False)


def register_view(request):
    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()

            # Сохраняем ФИО
            user.last_name = form.cleaned_data.get("last_name", "")
            user.first_name = form.cleaned_data.get("first_name", "")
            user.save()

            # Сохраняем профиль
            profile = user.profile
            profile.phone = form.cleaned_data.get("phone", "")
            profile.position = form.cleaned_data.get("position", "")
            profile.patronymic = form.cleaned_data.get("patronymic", "")
            profile.yuid_name = form.cleaned_data.get("yuid_name", "")

            inn = form.cleaned_data.get("inn")
            if inn:
                org = Organization.objects.filter(inn=inn).first()
                if org:
                    profile.organization = org
            profile.save()

            login(request, user)
            messages.success(
                request, f"Добро пожаловать, {user.first_name} {user.last_name}!"
            )
            return redirect("home")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = RegistrationForm()

    return render(request, "register.html", {"form": form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect("home")
    
    if request.method == "POST":
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")
            user = authenticate(username=username, password=password)
            if user is not None:
                profile, created = UserProfile.objects.get_or_create(user=user)
                login(request, user)
                messages.success(request, f"С возвращением, {username}!")
                if user.is_superuser:
                    return redirect("admin_dashboard")
                else:
                    return redirect("home")
        messages.error(request, "Неверное имя пользователя или пароль")
    else:
        form = LoginForm()

    return render(request, "login.html", {"form": form})


def logout_view(request):
    logout(request)
    messages.info(request, "Вы вышли из системы")
    return redirect("login")


@login_required
def profile_view(request):
    profile = request.user.profile
    user = request.user

    if request.method == "POST":
        # Сохраняем ФИО
        user.last_name = request.POST.get("last_name", "")
        user.first_name = request.POST.get("first_name", "")
        user.save()

        # Сохраняем профиль
        profile.phone = request.POST.get("phone", "")
        profile.position = request.POST.get("position", "")
        profile.patronymic = request.POST.get("patronymic", "")
        profile.yuid_name = request.POST.get("yuid_name", "")
        profile.save()

        messages.success(request, "Профиль успешно обновлен")
        return redirect("profile")

    return render(request, "profile.html", {"profile": profile, "user": user})


@login_required
def dashboard(request):
    profile = request.user.profile

    # Центр Свердловской области
    SVERDLOVSK_CENTER = [56.8439, 60.6524]

    # Список районов
    districts_list = [
        'Автовокзал', 'Академический', 'Березит', 'Ботанический', 'Верхнемакарово',
        'ВИЗ', 'Вокзальный', 'Втузгородок', 'Горный Щит', 'Елизавет', 'ЖБИ',
        'Завокзальный', 'Заречный', 'Изоплит', 'Исток', 'Калиновский', 'Карасьеозёрский',
        'Кольцово', 'Компрессорный', 'Лечебный', 'Малый Исток', 'Медный', 'Нижне-Исетский',
        'Новая Сортировка', 'Палкинский Торфяник', 'Палкино', 'Палникс', 'Парковый',
        'Пионерский', 'Полеводство', 'Птицефабрика', 'Рудный', 'Садовый', 'Северка',
        'Семь ключей', 'Сибирский', 'Синие Камни', 'Совхоз', 'Солнечный', 'Старая Сортировка',
        'Сулимовский', 'Технопарк', 'Уктус', 'УНЦ', 'Уралмаш', 'Химмаш', 'Центральный',
        'Чермет', 'Чусовское Озеро', 'Шабровский', 'Шарташ', 'Шарташский р-к',
        'Широкая Речка', 'Шувакиш', 'Эльмаш', 'Юго-Западный', 'Веер'
    ]

    # Список годов для фильтра ДТП (из базы)
    accident_years = list(Accident.objects.dates('date', 'year').values_list('date__year', flat=True).distinct().order_by('-date__year'))
    if not accident_years:
        accident_years = [datetime.now().year]

    # Права доступа
    is_propaganda = request.user.is_superuser or profile.user_type == "propaganda"

    # Мероприятия: ГИБДД видит все, обычный пользователь — только свои
    if is_propaganda:
        events = Event.objects.all().select_related("organization", "created_by")
    else:
        events = Event.objects.filter(organization=profile.organization).select_related(
            "organization", "created_by"
        )
    
   # Сортировка (для всех пользователей)
    sort = request.GET.get('sort', '-created_at')
    allowed_sorts = ['created_at', '-created_at', 'date', '-date']
    if sort in allowed_sorts:
        events = events.order_by(sort)
    else:
        events = events.order_by('-created_at')

    # Пагинация: 15 записей на страницу
    paginator = Paginator(events, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # ========== Формируем данные для карты (мероприятия) — ОДИН РАЗ ==========
    events_for_map = []
    for event in events:
        if event.latitude and event.longitude:
            events_for_map.append({
                "id": event.id,
                "title": event.title,
                "latitude": float(event.latitude),
                "longitude": float(event.longitude),
                "address": event.location_address or "",
                "status": event.status,
                "event_type": event.get_event_type_display(),
                "date": event.date.strftime("%d.%m.%Y") if event.date else "",
            })

    # ========== Формируем данные для организаций (школ) — с учётом прав ==========
    schools_for_map = []
    organizations = Organization.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)

    for org in organizations:
        # ГИБДД видит все организации, обычный пользователь — только свою
        if is_propaganda or (profile.organization and profile.organization.id == org.id):
            schools_for_map.append({
                "id": org.id,
                "name": org.name,
                "latitude": float(org.latitude),
                "longitude": float(org.longitude),
                "address": org.address or ""
            })

    # ========== Обработка POST-запроса (добавление мероприятия) ==========
    if request.method == "POST":
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save(commit=False)
            if not profile.organization:
                messages.error(request, "У вашего профиля не привязана организация")
                return redirect("home")
            event.organization = profile.organization
            event.created_by = request.user
            event.district = request.POST.get('district', '')
            event.save()

            files = request.FILES.getlist('report_files')
            for f in files:
                ReportFile.objects.create(
                    event=event,
                    file=f,
                    filename=f.name
                )

            messages.success(request, "Мероприятие успешно добавлено!")
            return redirect("home")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return redirect("home")
    else:
        form = EventForm()

    # ========== Все теги (для фильтра) ==========
    all_tags = sorted(set(
        tag.strip()
        for event in Event.objects.exclude(tags__isnull=True).exclude(tags='')
        for tag in event.tags.split(',')
        if tag.strip()
    ))

    return render(
        request,
        "dashboard.html",
        {
            "events": page_obj,
            "page_obj": page_obj,
            "events_for_map": json.dumps(events_for_map, ensure_ascii=False),
            "map_center": SVERDLOVSK_CENTER,
            "form": form,
            "YANDEX_API_KEY": settings.YANDEX_API_KEY,
            "is_propaganda": is_propaganda,
            "all_tags": all_tags,
            "districts_list": districts_list,
            "accident_years": accident_years,
            "schools_for_map": json.dumps(schools_for_map, ensure_ascii=False),
            "all_schools": Organization.objects.filter(
                latitude__isnull=False, longitude__isnull=False
            ).exclude(latitude=0, longitude=0) if is_propaganda else Organization.objects.none(),
        },
    )

@login_required
def delete_event(request, event_id):
    try:
        event = Event.objects.get(id=event_id)
        if request.user.is_superuser or request.user.profile.user_type == "propaganda":
            event.delete()
            messages.success(request, "Мероприятие удалено")
        elif event.organization == request.user.profile.organization:
            event.delete()
            messages.success(request, "Мероприятие удалено")
        else:
            messages.error(request, "Нет прав для удаления")
    except Event.DoesNotExist:
        messages.error(request, "Мероприятие не найдено")
    return redirect("home")


@login_required
def get_event_json(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Получаем список файлов
    files_list = []
    for f in event.report_files.all():
        files_list.append({
            'id': f.id,
            'name': f.filename,
            'url': f.file.url
        })

    return JsonResponse({
        "id": event.id,
        "title": event.title,
        "date": event.date.strftime("%Y-%m-%d") if event.date else "",
        "time": event.time.strftime("%H:%M") if event.time else "",
        "status": event.status,
        "event_type": event.event_type,
        "location_type": event.location_type,
        "location_address": event.location_address or "",
        "latitude": event.latitude,
        "longitude": event.longitude,
        "description": event.description or "",
        "tags": event.tags or "",
        "responsible_person": event.responsible_person or "",
        "report_link": event.report_link or "",
        "report_files": files_list,
        "district": event.district or "",
    })


@csrf_exempt
@login_required
def update_event_json(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    if request.method == "POST":
        # Проверяем, есть ли файлы в запросе
        if request.FILES.getlist('report_files'):
            files = request.FILES.getlist('report_files')
            from .models import ReportFile
            for f in files:
                ReportFile.objects.create(
                    event=event,
                    file=f,
                    filename=f.name
                )
            return JsonResponse({"success": True, "files_added": len(files)})
        
        # Обработка обычных полей (JSON)
        try:
            data = json.loads(request.body)
            
            event.title = data.get("title", event.title)
            event.date = datetime.strptime(data.get("date", ""), "%Y-%m-%d").date() if data.get("date") else event.date
            event.time = data.get("time") if data.get("time") else None
            event.event_type = data.get("event_type", event.event_type)
            new_status = data.get("status", event.status)
            # Ограничения для обычного пользователя
            if not request.user.is_superuser and request.user.profile.user_type != 'propaganda':
                # Нельзя поставить "conducted" если не было "approved"
                if event.status != 'approved' and new_status == 'conducted':
                    return JsonResponse({"error": "Нельзя поставить статус 'Проведено' без подтверждения ГИБДД"}, status=400)
                # Если статус "approved", можно только "conducted" или оставить как есть
                if event.status == 'approved' and new_status not in ['conducted', 'approved']:
                    return JsonResponse({"error": "Можно изменить только на 'Проведено'"}, status=400)

            event.status = new_status
            event.location_type = data.get("location_type", event.location_type)
            event.location_address = data.get("location_address", event.location_address)
            event.description = data.get("description", event.description)
            event.latitude = data.get("latitude", event.latitude)
            event.longitude = data.get("longitude", event.longitude)
            event.tags = data.get("tags", event.tags)
            event.responsible_person = data.get("responsible_person", event.responsible_person)
            event.report_link = data.get("report_link", event.report_link)
            event.district = data.get("district", event.district)
            event.save()
            
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    
    return JsonResponse({"error": "Метод не разрешен"}, status=405)

@csrf_exempt
@login_required
def delete_report_file(request, file_id):
    """Удаление файла отчета"""
    if request.method == "POST":
        try:
            file = get_object_or_404(ReportFile, id=file_id)
            # Проверяем права: пользователь может удалить файл только если он создатель мероприятия
            if request.user == file.event.created_by or request.user.is_superuser:
                file.delete()
                return JsonResponse({"success": True})
            return JsonResponse({"error": "Нет прав"}, status=403)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Метод не разрешен"}, status=405)


@login_required
def map_view(request):
    events = Event.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)

    accidents = Accident.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)

    organizations = Organization.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)

    events_data = [
        {
            "id": e.id,
            "title": e.title,
            "description": e.description or "",
            "location_address": e.location_address or "",
            "latitude": e.latitude,
            "longitude": e.longitude,
            "date": e.date.strftime("%d.%m.%Y") if e.date else "",
        }
        for e in events
    ]

    accidents_data = [
        {
            "id": a.id,
            "title": a.title,
            "description": a.description or "",
            "latitude": a.latitude,
            "longitude": a.longitude,
            "date": a.date.strftime("%d.%m.%Y %H:%M") if a.date else "",
        }
        for a in accidents
    ]

    organizations_data = [
        {
            "id": o.id,
            "name": o.name,
            "address": o.address or "",
            "latitude": o.latitude,
            "longitude": o.longitude,
        }
        for o in organizations
    ]

    # Статистика по районам (мероприятия + ДТП)
    from django.db.models import Count
    # Список районов
    districts_list = [
        'Автовокзал', 'Академический', 'Березит', 'Ботанический', 'Верхнемакарово',
        'ВИЗ', 'Вокзальный', 'Втузгородок', 'Горный Щит', 'Елизавет', 'ЖБИ',
        'Завокзальный', 'Заречный', 'Изоплит', 'Исток', 'Калиновский', 'Карасьеозёрский',
        'Кольцово', 'Компрессорный', 'Лечебный', 'Малый Исток', 'Медный', 'Нижне-Исетский',
        'Новая Сортировка', 'Палкинский Торфяник', 'Палкино', 'Палникс', 'Парковый',
        'Пионерский', 'Полеводство', 'Птицефабрика', 'Рудный', 'Садовый', 'Северка',
        'Семь ключей', 'Сибирский', 'Синие Камни', 'Совхоз', 'Солнечный', 'Старая Сортировка',
        'Сулимовский', 'Технопарк', 'Уктус', 'УНЦ', 'Уралмаш', 'Химмаш', 'Центральный',
        'Чермет', 'Чусовское Озеро', 'Шабровский', 'Шарташ', 'Шарташский р-к',
        'Широкая Речка', 'Шувакиш', 'Эльмаш', 'Юго-Западный', 'Веер'
    ]

        

    district_stats = []
    for district in districts_list:
        events = Event.objects.filter(district=district)
        accidents = Accident.objects.filter(district=district)
        event_types = events.values('event_type').annotate(cnt=Count('id'))
        types_str = ', '.join([f"{e['event_type']}({e['cnt']})" for e in event_types]) or '—'
        
        district_stats.append({
            'district': district,
            'events_count': events.count(),
            'event_types': types_str,
            'accidents_count': accidents.count(),
        })

    context = {
        "events_json": json.dumps(events_data, ensure_ascii=False),
        "accidents_json": json.dumps(accidents_data, ensure_ascii=False),
        "organizations_json": json.dumps(organizations_data, ensure_ascii=False),
        "total_events": events.count(),
        "total_accidents": accidents.count(),
        "total_organizations": organizations.count(),
        "YANDEX_API_KEY": settings.YANDEX_API_KEY,
        'district_stats': district_stats, 
    }

    return render(request, "map.html", context)


def debug_session(request):
    events = request.session.get(SESSION_KEY, [])
    return HttpResponse(f"""
    <html>
    <body style="font-family: monospace; padding: 20px; background: #f0f0f0;">
        <h1>Отладка сессии</h1>
        <p>Ключ сессии: {SESSION_KEY}</p>
        <p>Количество мероприятий: {len(events)}</p>
        <hr>
        <a href="/">Вернуться на главную</a>
    </body>
    </html>
    """)

import joblib
from pathlib import Path

def ml_recommendation(request):
    """ML рекомендация по району с прогнозом"""
    district = request.GET.get('district', '')
    
    if not district:
        return JsonResponse({'error': 'Район не указан'}, status=400)
    
    from events.models import Accident, Event
    
    # Статистика по району
    accidents = Accident.objects.filter(district=district)
    total_accidents = accidents.count()
    fatal = accidents.filter(severity='fatal').count()
    heavy = accidents.filter(severity='heavy').count()
    
    # Мероприятия в районе
    events_count = Event.objects.filter(district=district).count()
    
    # Расчёт уровня риска
    if total_accidents > 15:
        risk_level = 'Высокий'
        recommended_event = 'Акция / Рейд'
        reason = f'В районе зафиксировано {total_accidents} ДТП, из них {fatal} смертельных'
    elif total_accidents > 5:
        risk_level = 'Средний'
        recommended_event = 'Практическое занятие'
        reason = f'В районе зафиксировано {total_accidents} ДТП'
    else:
        risk_level = 'Низкий'
        recommended_event = 'Лекция / Беседа'
        reason = f'В районе зафиксировано {total_accidents} ДТП'
    
    # Прогноз тяжести следующего ДТП
    if heavy > 3 or fatal > 1:
        severity_prediction = 'Тяжёлое или смертельное'
    elif heavy > 0:
        severity_prediction = 'Средней тяжести'
    else:
        severity_prediction = 'Лёгкое'
    
    # Рекомендация
    if risk_level == 'Высокий':
        recommendation = f'{recommended_event}. {reason}. Прогнозируемая тяжесть следующего ДТП: {severity_prediction}. Требуется усиленная профилактика.'
    elif risk_level == 'Средний':
        recommendation = f'{recommended_event}. {reason}. Прогнозируемая тяжесть следующего ДТП: {severity_prediction}. Рекомендуется регулярная профилактика.'
    else:
        recommendation = f'{recommended_event}. {reason}. Прогнозируемая тяжесть следующего ДТП: {severity_prediction}. Достаточно поддерживающих мероприятий.'
    
    return JsonResponse({
        'success': True,
        'district': district,
        'risk_level': risk_level,
        'recommended_event': recommended_event,
        'severity_prediction': severity_prediction,
        'total_accidents': total_accidents,
        'recommendation': recommendation
    })

@login_required
def map_all_view(request):
    """Карта с мероприятиями и ДТП (два слоя)"""
    
    # Мероприятия с координатами
    events = Event.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)
    
    # ДТП с координатами
    accidents = Accident.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)
    
    events_data = []
    for e in events:
        events_data.append({
            'id': e.id,
            'title': e.title,
            'latitude': e.latitude,
            'longitude': e.longitude,
            'address': e.location_address or '',
            'status': e.status,
            'event_type': e.get_event_type_display(),
            'date': e.date.strftime('%d.%m.%Y') if e.date else '',
        })
    
    accidents_data = []
    for a in accidents:
        accidents_data.append({
            'id': a.id,
            'title': a.title,
            'latitude': a.latitude,
            'longitude': a.longitude,
            'address': a.location_address or '',
            'severity': a.severity,
            'date': a.date.strftime('%d.%m.%Y %H:%M') if a.date else '',
            'description': a.description or '',
        })
    
    context = {
        'events_json': json.dumps(events_data, ensure_ascii=False),
        'accidents_json': json.dumps(accidents_data, ensure_ascii=False),
        'total_events': events.count(),
        'total_accidents': accidents.count(),
        'YANDEX_API_KEY': settings.YANDEX_API_KEY,
    }
    
    return render(request, 'map_with_accidents.html', context)


def geocode_address(address):
    """Получить координаты по адресу через Яндекс.Геокодер"""
    if not address:
        return None, None
    
    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {
        'apikey': settings.YANDEX_API_KEY,
        'geocode': address,
        'format': 'json',
        'results': 1
    }
    
    try:
        response = requests.get(url, params=params, timeout=5)
        data = response.json()
        
        # Проверяем, есть ли результаты
        if not data.get('response', {}).get('GeoObjectCollection', {}).get('featureMember'):
            return None, None
        
        point = data['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']
        lng, lat = point.split()
        return float(lat), float(lng)
    except Exception as e:
        print(f"Ошибка геокодирования {address}: {e}")
        return None, None

@login_required
def upload_accidents(request):
    """Страница загрузки данных (Excel или GeoJSON)"""
    is_propaganda = request.user.is_superuser or request.user.profile.user_type == 'propaganda'
    
    # Загружаем границы районов Екатеринбурга
    districts_data = None
    geojson_path = os.path.join(settings.BASE_DIR, 'static/data/ekb-districts.json')
    try:
        with open(geojson_path, 'r', encoding='utf-8') as f:
            districts_data = json.load(f)
        print(f"Загружены границы районов из {geojson_path}")
    except Exception as e:
        print(f"Ошибка загрузки границ районов: {e}")
    
    def is_point_in_polygon(point, polygon_coords):
        """point: (долгота, широта), polygon_coords: [(широта, долгота), ...]"""
        lng, lat = point
        inside = False
        for i in range(len(polygon_coords)):
            x1, y1 = polygon_coords[i]
            x2, y2 = polygon_coords[(i + 1) % len(polygon_coords)]
            if ((y1 > lat) != (y2 > lat)) and (lng < (x2 - x1) * (lat - y1) / (y2 - y1) + x1):
                inside = not inside
        return inside
    
    def get_district_by_coords(lng, lat):
        """Определяет район по координатам (долгота, широта) - исправленная версия"""
        load_districts_data()
        
        if not DISTRICTS_DATA:
            return ''
        if lng is None or lat is None:
            return ''
        
        def point_in_polygon(point, polygon_coords):
            x, y = point  # x = долгота, y = широта
            inside = False
            for i in range(len(polygon_coords)):
                x1, y1 = polygon_coords[i]
                x2, y2 = polygon_coords[(i + 1) % len(polygon_coords)]
                if ((y1 > y) != (y2 > y)) and (x < (x2 - x1) * (y - y1) / (y2 - y1) + x1):
                    inside = not inside
            return inside
        
        point = (lng, lat)
        
        for feature in DISTRICTS_DATA['features']:
            name = feature['properties']['name']
            geometry = feature['geometry']
            
            if geometry['type'] == 'Polygon':
                poly_coords = geometry['coordinates'][0]
                if point_in_polygon(point, poly_coords):
                    print(f"Попадание в район: {name} для точки {lng}, {lat}")
                    return name
            elif geometry['type'] == 'MultiPolygon':
                for polygon in geometry['coordinates']:
                    poly_coords = polygon[0]
                    if point_in_polygon(point, poly_coords):
                        print(f"Попадание в район: {name} для точки {lng}, {lat}")
                        return name
        
        print(f"Точка {lng}, {lat} не попала ни в один полигон")
        return ''
    
    if request.method == 'POST':
        upload_type = request.POST.get('upload_type')
        uploaded_file = request.FILES.get('excel_file')
        
        if not uploaded_file:
            messages.error(request, 'Выберите файл')
            return redirect('upload_accidents')
        
        filename = uploaded_file.name.lower()
        
        # ==================== GEOJSON ЗАГРУЗКА ====================
        if filename.endswith(('.geojson', '.json')):
            if upload_type != 'accidents':
                messages.error(request, 'GeoJSON можно загружать только как ДТП')
                return redirect('upload_accidents')
            
            if not is_propaganda:
                messages.error(request, 'Нет прав для загрузки ДТП')
                return redirect('upload_accidents')
            
            try:
                data = json.load(uploaded_file)
                features = data.get('features', [])
                added = 0
                
                for feature in features:
                    props = feature.get('properties', {})
                    geometry = feature.get('geometry', {})
                    coords = geometry.get('coordinates', [])
                    
                    # Координаты: [долгота, широта]
                    if len(coords) >= 2:
                        longitude = coords[0]
                        latitude = coords[1]
                    else:
                        latitude = None
                        longitude = None
                    
                    title = props.get('category', 'ДТП') or 'ДТП'
                    address = props.get('address', '')
                    
                    # Собираем все ключевые слова в описание
                    tags = props.get('tags', [])
                    if isinstance(tags, list):
                        tags_text = ', '.join(str(t) for t in tags)
                    else:
                        tags_text = str(tags) if tags else ''
                    
                    light = props.get('light', '')
                    weather_list = props.get('weather', [])
                    if isinstance(weather_list, list):
                        weather_text = ', '.join(str(w) for w in weather_list)
                    else:
                        weather_text = str(weather_list) if weather_list else ''
                    
                    category_text = props.get('category', '')
                    nearby_list = props.get('nearby', [])
                    if isinstance(nearby_list, list):
                        nearby_text = ', '.join(str(n) for n in nearby_list)
                    else:
                        nearby_text = str(nearby_list) if nearby_list else ''
                    
                    road_conditions_list = props.get('road_conditions', [])
                    if isinstance(road_conditions_list, list):
                        road_text = ', '.join(str(r) for r in road_conditions_list)
                    else:
                        road_text = str(road_conditions_list) if road_conditions_list else ''
                    
                    # Собираем нарушения из всех участников
                    violations = []
                    for vehicle in props.get('vehicles', []):
                        for participant in vehicle.get('participants', []):
                            for v in participant.get('violations', []):
                                if v != 'Нет нарушений':
                                    violations.append(v)
                    violations_text = ', '.join(violations) if violations else ''
                    
                    # Объединяем всё в описание
                    description_parts = []
                    if tags_text:
                        description_parts.append(f"Теги: {tags_text}")
                    if light:
                        description_parts.append(f"Освещение: {light}")
                    if weather_text:
                        description_parts.append(f"Погода: {weather_text}")
                    if category_text:
                        description_parts.append(f"Категория: {category_text}")
                    if road_text:
                        description_parts.append(f"Дорожные условия: {road_text}")
                    if nearby_text:
                        description_parts.append(f"Объекты: {nearby_text}")
                    if violations_text:
                        description_parts.append(f"Нарушения: {violations_text}")
                    
                    description_text = '. '.join(description_parts)
                    
                    datetime_str = props.get('datetime', '')
                    
                    # Преобразуем тяжесть
                    severity_raw = props.get('severity', 'Легкий')
                    severity_map = {
                        'Легкий': 'light',
                        'Тяжёлый': 'heavy',
                        'С погибшими': 'fatal'
                    }
                    severity = severity_map.get(severity_raw, 'light')
                    
                    try:
                        accident_date = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
                    except:
                        accident_date = datetime.now()
                    
                    # Определяем район только по координатам
                    district = get_district_by_coords(longitude, latitude)
                    print(f"GEOJSON: координаты={longitude}, {latitude}, район={district}") 
                    
                    # Собираем теги одной строкой
                    all_tags_list = []
                    if tags_text: all_tags_list.append(tags_text)
                    if light: all_tags_list.append(light)
                    if weather_text: all_tags_list.append(weather_text)
                    if category_text: all_tags_list.append(category_text)
                    if road_text: all_tags_list.append(road_text)
                    if violations_text: all_tags_list.append(violations_text)
                    tags_field = ', '.join(all_tags_list)

                    Accident.objects.create(
                        title=title,
                        description=description_text,
                        date=accident_date,
                        location_address=address,
                        district=district,
                        latitude=latitude,
                        longitude=longitude,
                        severity=severity,
                        tags=tags_field
                    )
                    added += 1
                
                messages.success(request, f'Загружено {added} ДТП из GeoJSON')
                
                if added >= 100:
                    from .ml_utils import train_models
                    try:
                        train_models()
                    except:
                        pass
                
                return redirect('upload_accidents')
                
            except json.JSONDecodeError as e:
                messages.error(request, f'Ошибка парсинга GeoJSON: {str(e)}')
                return redirect('upload_accidents')
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
                return redirect('upload_accidents')
        
        # ==================== EXCEL ЗАГРУЗКА ====================
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                df = pd.read_excel(uploaded_file)
                
                # Общие маппинги
                status_map = {
                    'Запланировано': 'planned',
                    'Проведено': 'conducted',
                    'Отменено': 'cancelled'
                }
                type_map = {
                    'Лекция / Беседа': 'lecture',
                    'Практическое занятие': 'practical',
                    'Викторина / Конкурс': 'quiz',
                    'Родительское собрание': 'parent_meeting',
                    'Инструктаж': 'briefing',
                    'Акция / Рейд': 'action',
                    'Экскурсия': 'excursion',
                    'Тренинг': 'training',
                    'Иное': 'other'
                }
                severity_map = {
                    'лёгкое': 'light', 'легкое': 'light',
                    'среднее': 'medium', 'тяжёлое': 'heavy',
                    'тяжелое': 'heavy', 'смертельное': 'fatal'
                }
                
                added = 0
                
                for _, row in df.iterrows():
                    address = str(row.get('Адрес', '')).strip()
                    title = row.get('Название')
                    
                    # Геокодирование адреса
                    latitude = None
                    longitude = None
                    district = row.get('Район', '')
                    
                    if address:
                        lat, lng = geocode_address(address)
                        if lat and lng:
                            latitude = lat
                            longitude = lng
                            dist = get_district_by_coords(longitude, latitude)
                            if dist:
                                district = dist
                    
                    # Загрузка ДТП
                    if upload_type == 'accidents':
                        date_str = row.get('Дата')
                        if date_str and isinstance(date_str, str):
                            try:
                                date = datetime.strptime(date_str, '%d.%m.%Y')
                            except:
                                date = datetime.now()
                        elif hasattr(date_str, 'date'):
                            date = date_str
                        else:
                            date = datetime.now()
                        
                        severity = str(row.get('Тяжесть', 'light')).lower()
                        severity = severity_map.get(severity, 'light')
                        
                        Accident.objects.create(
                            title=title or f"ДТП на {address[:50]}" if address else "ДТП",
                            description=row.get('Описание', ''),
                            date=date,
                            location_address=address,
                            district=district,
                            severity=severity,
                            latitude=latitude,
                            longitude=longitude
                        )
                        added += 1
                    
                    # Загрузка мероприятий
                    elif upload_type == 'events':
                        if not title:
                            continue
                        
                        date_value = row.get('Дата')
                        if date_value:
                            if isinstance(date_value, str):
                                try:
                                    date = datetime.strptime(date_value, '%d.%m.%Y').date()
                                except:
                                    date = datetime.now().date()
                            elif hasattr(date_value, 'date'):
                                date = date_value.date()
                            else:
                                date = datetime.now().date()
                        else:
                            date = datetime.now().date()
                        
                        status_raw = row.get('Статус', 'Запланировано')
                        status = status_map.get(status_raw, 'planned')
                        
                        type_raw = row.get('Тип', 'Лекция / Беседа')
                        event_type = type_map.get(type_raw, 'lecture')
                        
                        Event.objects.create(
                            title=title,
                            date=date,
                            status=status,
                            event_type=event_type,
                            district=district,
                            location_address=address,
                            organization=profile.organization,
                            created_by=request.user,
                            responsible_person=row.get('Ответственное лицо', ''),
                            tags=row.get('Теги', ''),
                            latitude=latitude,
                            longitude=longitude
                        )
                        added += 1
                
                if upload_type == 'accidents':
                    messages.success(request, f'Загружено {added} ДТП')
                    
                    if added >= 100:
                        from .ml_utils import train_models
                        try:
                            train_models()
                        except:
                            pass
                else:
                    messages.success(request, f'Загружено {added} мероприятий')
                
                return redirect('upload_accidents')
                
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
                return redirect('upload_accidents')
        
        else:
            messages.error(request, 'Неподдерживаемый формат файла. Загрузите .xlsx, .xls, .geojson или .json')
            return redirect('upload_accidents')
    
    return render(request, 'upload_accidents.html', {'is_propaganda': is_propaganda})

# Авто-обучение если добавлено > 100 новых ДТП
from .ml_utils import train_models
import os

model_path = os.path.join(settings.BASE_DIR, 'ml_models', 'category_classifier.pkl')


def export_accidents_excel(request):
    """Экспорт ДТП в Excel"""
    if not request.user.is_superuser and request.user.profile.user_type != 'propaganda':
        messages.error(request, 'Нет прав для экспорта')
        return redirect('home')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ДТП"
    
    headers = ['ID', 'Название', 'Дата', 'Адрес', 'Район', 'Тяжесть', 'Описание']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    from .models import Accident
    accidents = Accident.objects.all().order_by('-date')
    for row, accident in enumerate(accidents, 2):
        ws.cell(row=row, column=1, value=accident.id)
        ws.cell(row=row, column=2, value=accident.title)
        ws.cell(row=row, column=3, value=accident.date.strftime('%d.%m.%Y') if accident.date else '')
        ws.cell(row=row, column=4, value=accident.location_address or '')
        ws.cell(row=row, column=5, value=accident.district or '')
        ws.cell(row=row, column=6, value=accident.get_severity_display())
        ws.cell(row=row, column=7, value=accident.description or '')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dtp.xlsx"'
    wb.save(response)
    return response


def export_events_excel(request):
    """Экспорт мероприятий в Excel"""
    if not request.user.is_superuser and request.user.profile.user_type != 'propaganda':
        messages.error(request, 'Нет прав для экспорта')
        return redirect('home')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мероприятия"
    
    headers = ['ID', 'Название', 'Дата', 'Время', 'Статус', 'Тип', 'Район', 'Адрес', 
               'Ответственное лицо', 'Теги', 'Ссылка на отчет', 'Дата создания']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    from .models import Event
    events = Event.objects.all().order_by('-date')
    for row, event in enumerate(events, 2):
        ws.cell(row=row, column=1, value=event.id)
        ws.cell(row=row, column=2, value=event.title)
        ws.cell(row=row, column=3, value=event.date.strftime('%d.%m.%Y') if event.date else '')
        ws.cell(row=row, column=4, value=event.time.strftime('%H:%M') if event.time else '')
        ws.cell(row=row, column=5, value=event.get_status_display())
        ws.cell(row=row, column=6, value=event.get_event_type_display())
        ws.cell(row=row, column=7, value=event.district or '')
        ws.cell(row=row, column=8, value=event.location_address or '')
        ws.cell(row=row, column=9, value=event.responsible_person or '')
        ws.cell(row=row, column=10, value=event.tags or '')
        ws.cell(row=row, column=11, value=event.report_link or '')
        ws.cell(row=row, column=12, value=event.created_at.strftime('%d.%m.%Y %H:%M') if event.created_at else '')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="meropriyatiya.xlsx"'
    wb.save(response)
    return response

@login_required
def plan_calendar(request):
    """Страница календаря мероприятий"""
    from .models import Event
    from datetime import datetime
    from calendar import monthrange
    import calendar
    
    profile = request.user.profile
    is_propaganda = request.user.is_superuser or profile.user_type == "propaganda"
    
    months_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    
    months_short_ru = {
        1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр',
        5: 'Май', 6: 'Июн', 7: 'Июл', 8: 'Авг',
        9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
    }
    
    current_year = datetime.now().year
    current_month = datetime.now().month
    today = datetime.now().day
    
    selected_year = int(request.GET.get('year', current_year))
    selected_month = int(request.GET.get('month', current_month))
    
    # Получаем мероприятия за выбранный месяц с учётом прав
    if is_propaganda:
        events = Event.objects.filter(
            date__year=selected_year,
            date__month=selected_month
        ).select_related('organization')
    else:
        events = Event.objects.filter(
            date__year=selected_year,
            date__month=selected_month,
            organization=profile.organization
        ).select_related('organization')
    
    events = events.order_by('date')
    
    # Группируем по дням
    events_by_day = {}
    for event in events:
        day = event.date.day
        if day not in events_by_day:
            events_by_day[day] = []
        events_by_day[day].append(event)
    
    # Строим календарную сетку
    first_weekday = calendar.weekday(selected_year, selected_month, 1)
    days_in_month = monthrange(selected_year, selected_month)[1]
    first_weekday = (first_weekday - 1) % 7
    
    calendar_grid = []
    week = []
    
    for _ in range(first_weekday):
        week.append({'day': 0, 'events': []})
    
    for day in range(1, days_in_month + 1):
        week.append({'day': day, 'events': events_by_day.get(day, [])})
        if len(week) == 7:
            calendar_grid.append(week)
            week = []
    
    if week:
        while len(week) < 7:
            week.append({'day': 0, 'events': []})
        calendar_grid.append(week)
    
    # Статистика
    total_events = len(events)
    event_types_count = {}
    status_count = {'planned': 0, 'conducted': 0, 'cancelled': 0, 'approved': 0}
    
    for event in events:
        type_name = event.get_event_type_display()
        event_types_count[type_name] = event_types_count.get(type_name, 0) + 1
        
        if event.status == 'planned':
            status_count['planned'] += 1
        elif event.status == 'conducted':
            status_count['conducted'] += 1
        elif event.status == 'cancelled':
            status_count['cancelled'] += 1
        elif event.status == 'approved':
            status_count['approved'] += 1
    
    # Существующие годы с мероприятиями
    existing_years = list(Event.objects.dates('date', 'year').values_list('date__year', flat=True).distinct())
    if not existing_years:
        existing_years = [current_year]
    
    # Навигация: только по годам с мероприятиями
    if selected_month > 1:
        prev_year = selected_year
        prev_month = selected_month - 1
    else:
        prev_month = 12
        prev_year = selected_year
        for y in sorted(existing_years, reverse=True):
            if y < selected_year:
                prev_year = y
                break
    
    if selected_month < 12:
        next_year = selected_year
        next_month = selected_month + 1
    else:
        next_month = 1
        next_year = selected_year
        for y in sorted(existing_years):
            if y > selected_year:
                next_year = y
                break
    
    months = [(i, months_short_ru[i]) for i in range(1, 13)]
    
    context = {
        'selected_year': selected_year,
        'selected_month': selected_month,
        'month_name_ru': months_ru[selected_month],
        'calendar_grid': calendar_grid,
        'total_events': total_events,
        'event_types_count': event_types_count,
        'status_count': status_count,
        'years': sorted(existing_years, reverse=True),
        'months': months,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'prev_month_name': months_ru[prev_month],
        'next_year': next_year,
        'next_month': next_month,
        'next_month_name': months_ru[next_month],
        'today': today,
        'current_month': current_month,
        'current_year': current_year,
        'is_propaganda': is_propaganda,
        'existing_years': sorted(existing_years, reverse=True),
    }
    return render(request, 'plan_calendar.html', context)

def register_by_invite(request, token):
    """Регистрация сотрудника ГИБДД по токену-приглашению"""
    try:
        invitation = Invitation.objects.get(token=token)
    except Invitation.DoesNotExist:
        messages.error(request, 'Неверная или устаревшая ссылка приглашения')
        return redirect('login')
    
    if not invitation.is_valid():
        messages.error(request, 'Ссылка приглашения уже использована или исчерпан лимит')
        return redirect('login')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        email = request.POST.get('email')
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        patronymic = request.POST.get('patronymic')
        phone = request.POST.get('phone')
        position = request.POST.get('position')
        
        if password1 != password2:
            messages.error(request, 'Пароли не совпадают')
            return redirect('register_by_invite', token=token)
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Пользователь с таким логином уже существует')
            return redirect('register_by_invite', token=token)
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Пользователь с таким email уже существует')
            return redirect('register_by_invite', token=token)
        
        # Создаём пользователя
        user = User.objects.create_user(
            username=username,
            password=password1,
            email=email,
            last_name=last_name,
            first_name=first_name
        )
        
        # Получаем или создаём профиль и принудительно обновляем
        profile, created = UserProfile.objects.get_or_create(user=user)
        profile.user_type = 'propaganda'
        profile.organization = None
        profile.patronymic = patronymic if patronymic else ''
        profile.phone = phone if phone else ''
        profile.position = position if position else ''
        profile.yuid_name = ''
        profile.save()
        
        # Принудительно обновляем через update (гарантия)
        UserProfile.objects.filter(user=user).update(
            user_type='propaganda',
            organization=None,
            patronymic=patronymic if patronymic else '',
            phone=phone if phone else '',
            position=position if position else '',
            yuid_name=''
        )
        
        # Увеличиваем счётчик использований токена
        invitation.used_count += 1
        invitation.save()
        
        login(request, user)
        messages.success(request, f'Добро пожаловать, {first_name} {last_name}!')
        return redirect('home')
    
    return render(request, 'register_invite.html', {'token': token})

from datetime import datetime
from django.db.models import Q

def accidents_by_month(request):
    """Возвращает ДТП за указанный год и месяц"""
    try:
        year = int(request.GET.get('year', datetime.now().year))
        month = int(request.GET.get('month', datetime.now().month))
    except (ValueError, TypeError):
        year = datetime.now().year
        month = datetime.now().month
    
    # Фильтруем ДТП по году и месяцу
    accidents = Accident.objects.filter(
        latitude__isnull=False,
        longitude__isnull=False,
        date__year=year,
        date__month=month
    ).exclude(latitude=0, longitude=0)
    
    # Ограничиваем количество для производительности
    accidents = accidents[:1000]
    
    result = []
    severity_map = {
        'light': 'Лёгкое',
        'medium': 'Среднее',
        'heavy': 'Тяжёлое',
        'fatal': 'Смертельное'
    }
    
    for a in accidents:
        result.append({
            'lat': a.latitude,
            'lng': a.longitude,
            'title': a.title,
            'description': a.description or '',
            'date': a.date.strftime('%d.%m.%Y'),
            'address': a.location_address or '',
            'severity': a.severity,
            'severity_text': severity_map.get(a.severity, 'Неизвестно')
        })
    
    return JsonResponse(result, safe=False)

def get_clustered_accidents(request):
    """Возвращает сгруппированные ДТП только в видимой области"""
    from django.db.models import Count
    from django.db import connection
    
    try:
        zoom = int(request.GET.get('zoom', 9))
        bbox = request.GET.get('bbox', '')
    except (ValueError, TypeError):
        zoom = 9
    
    # Парсим границы карты
    bounds = None
    if bbox:
        try:
            parts = bbox.split(',')
            if len(parts) == 4:
                bounds = {
                    'min_lng': float(parts[0]),
                    'min_lat': float(parts[1]),
                    'max_lng': float(parts[2]),
                    'max_lat': float(parts[3])
                }
        except:
            pass
    
    if not bounds:
        return JsonResponse([], safe=False)
    
    # Точность округления зависит от масштаба
    if zoom <= 9:
        precision = 0.05
    elif zoom <= 11:
        precision = 0.02
    elif zoom <= 13:
        precision = 0.01
    else:
        precision = 0.005
    
    # SQL запрос с округлением координат
    query_sql = """
        SELECT 
            ROUND(latitude / %s) * %s as lat_rounded,
            ROUND(longitude / %s) * %s as lng_rounded,
            severity,
            COUNT(*) as count
        FROM events_accident
        WHERE latitude IS NOT NULL 
          AND longitude IS NOT NULL
          AND latitude >= %s
          AND latitude <= %s
          AND longitude >= %s
          AND longitude <= %s
        GROUP BY lat_rounded, lng_rounded, severity
        ORDER BY count DESC
        LIMIT 200
    """
    
    with connection.cursor() as cursor:
        cursor.execute(query_sql, [
            precision, precision,
            precision, precision,
            bounds['min_lat'], bounds['max_lat'],
            bounds['min_lng'], bounds['max_lng']
        ])
        rows = cursor.fetchall()
    
    result = []
    for row in rows:
        result.append({
            'lat': float(row[0]),
            'lng': float(row[1]),
            'severity': row[2],
            'count': row[3]
        })
    
    return JsonResponse(result, safe=False)

@csrf_exempt
def accidents_by_months(request):
    """Возвращает ДТП за указанный год, месяцы и типы тяжести"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = int(data.get('year', datetime.now().year))
            months = data.get('months', [])
            severities = data.get('severities', ['light', 'heavy', 'fatal'])
        except:
            return JsonResponse({'error': 'Invalid data'}, status=400)
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Фильтруем ДТП по году, месяцам и тяжести
    accidents = Accident.objects.filter(
        latitude__isnull=False,
        longitude__isnull=False,
        date__year=year,
        date__month__in=months,
        severity__in=severities
    ).exclude(latitude=0, longitude=0)
    
    # Ограничиваем количество для производительности
    accidents = accidents[:2000]
    
    result = []
    severity_map = {
        'light': 'Лёгкое',
        'heavy': 'Тяжёлое',
        'fatal': 'Смертельное'
    }
    
    for a in accidents:
        result.append({
            'lat': a.latitude,
            'lng': a.longitude,
            'title': a.title,
            'description': a.description or '',
            'date': a.date.strftime('%d.%m.%Y'),
            'address': a.location_address or '',
            'severity': a.severity,
            'severity_text': severity_map.get(a.severity, 'Неизвестно')
        })
    
    return JsonResponse(result, safe=False)

@csrf_exempt
@login_required
def approve_event(request, event_id):
    """Подтверждение мероприятия сотрудником ГИБДД"""
    if request.user.profile.user_type != 'propaganda' and not request.user.is_superuser:
        return JsonResponse({'error': 'Нет прав'}, status=403)
    
    event = get_object_or_404(Event, id=event_id)
    # Если мероприятие проведено — редактирование запрещено для обычных пользователей
    if event.status == 'conducted' and not request.user.is_superuser and request.user.profile.user_type != 'propaganda':
        return JsonResponse({"error": "Проведённое мероприятие нельзя редактировать"}, status=403)
    
    if event.status != 'planned':
        return JsonResponse({'error': 'Можно подтверждать только запланированные мероприятия'}, status=400)
    
    event.status = 'approved'
    event.gibdd_approved = True
    event.save()
    
    return JsonResponse({'success': True})

import json
from django.http import JsonResponse
from .models import Accident

def update_accidents_districts(request):
    """Обновить районы для ДТП на основе GeoJSON"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Нет прав'}, status=403)
    
    # Загружаем границы районов
    import os
    from django.conf import settings
    
    geojson_path = os.path.join(settings.BASE_DIR, 'static/data/ekb-districts.json')
    with open(geojson_path, 'r', encoding='utf-8') as f:
        districts_data = json.load(f)
    
    # Функция проверки точки внутри полигона
    def is_point_in_polygon(point, polygon_coords):
        lat, lng = point
        inside = False
        for i in range(len(polygon_coords)):
            x1, y1 = polygon_coords[i]
            x2, y2 = polygon_coords[(i + 1) % len(polygon_coords)]
            if ((y1 > lng) != (y2 > lng)) and (lat < (x2 - x1) * (lng - y1) / (y2 - y1) + x1):
                inside = not inside
        return inside
    
    # Получаем все ДТП без района
    accidents = Accident.objects.filter(district__isnull=True) | Accident.objects.filter(district='')
    updated = 0
    
    for accident in accidents:
        if accident.latitude and accident.longitude:
            point = (accident.longitude, accident.latitude)  # (долгота, широта)
            found_district = None
            
            for feature in districts_data['features']:
                name = feature['properties']['name']
                geometry = feature['geometry']
                
                if geometry['type'] == 'Polygon':
                    coords = [[c[1], c[0]] for c in geometry['coordinates'][0]]  # [широта, долгота]
                    if is_point_in_polygon(point, coords):
                        found_district = name
                        break
                        
                elif geometry['type'] == 'MultiPolygon':
                    for polygon in geometry['coordinates']:
                        coords = [[c[1], c[0]] for c in polygon[0]]
                        if is_point_in_polygon(point, coords):
                            found_district = name
                            break
                    if found_district:
                        break
            
            if found_district:
                accident.district = found_district
                accident.save()
                updated += 1
    
    return JsonResponse({'updated': updated, 'total': accidents.count()})

@login_required
def train_model_view(request):
    """Страница обучения модели (только для админа)"""
    if not request.user.is_superuser:
        messages.error(request, 'Нет доступа')
        return redirect('home')
    
    from .ml_utils import train_models
    import os
    
    if request.method == 'POST':
        try:
            accuracy = train_models()
            messages.success(request, f'Модель обучена! Точность: {accuracy:.2%}')
        except Exception as e:
            messages.error(request, f'Ошибка обучения: {e}')
        return redirect('train_model')
    
    # Проверяем, обучена ли модель
    model_exists = os.path.exists(os.path.join(settings.BASE_DIR, 'ml_models', 'category_classifier.pkl'))
    
    return render(request, 'train_model.html', {
        'model_exists': model_exists,
        'is_propaganda': request.user.is_superuser or request.user.profile.user_type == 'propaganda',
    })


@login_required
def school_analysis(request, school_id):
    """Страница анализа ДТП для школы"""
    from .ml_utils import analyze_school
    from datetime import datetime
    
    school = get_object_or_404(Organization, id=school_id)
    profile = request.user.profile
    is_propaganda = request.user.is_superuser or profile.user_type == 'propaganda'
    
    # Доступ: ГИБДД или сама школа
    if not is_propaganda and profile.organization_id != school.id:
        messages.error(request, 'Нет доступа')
        return redirect('home')
    
    year = request.GET.get('year')
    if year:
        try:
            year = int(year)
        except:
            year = datetime.now().year
    else:
        year = datetime.now().year
    
    analysis = analyze_school(school_id, year=year)
    
    available_years = list(range(2015, datetime.now().year + 1))
    available_years.reverse()
    
    context = {
        'school': school,
        'analysis': analysis,
        'selected_year': year,
        'available_years': available_years,
        'is_propaganda': is_propaganda,
    }
    
    return render(request, 'school_analysis.html', context)


def gibdd_summary_view(request):
    from .ml_utils import get_all_schools_summary
    from datetime import datetime
    
    profile = request.user.profile
    is_propaganda = request.user.is_superuser or profile.user_type == 'propaganda'
    
    if not is_propaganda:
        messages.error(request, 'Нет доступа')
        return redirect('home')
    
    year = request.GET.get('year')
    if year:
        try:
            year = int(year)
        except:
            year = datetime.now().year
    else:
        year = datetime.now().year
    
    summary = get_all_schools_summary(year=year)
    available_years = list(range(2015, datetime.now().year + 1))
    available_years.reverse()
    
    # Статистика
    total_dtp = sum(s['total_accidents'] for s in summary)
    dangerous_schools = sum(1 for s in summary if s['danger_level'] == 'Высокий')
    
    context = {
        'summary': summary,
        'selected_year': year,
        'available_years': available_years,
        'is_propaganda': is_propaganda,
        'total_dtp': total_dtp,
        'dangerous_schools': dangerous_schools,
    }
    
    return render(request, 'gibdd_summary.html', context)

def api_school_analysis(request):
    """API для анализа школы (возвращает JSON)"""
    from .ml_utils import analyze_school
    
    school_id = request.GET.get('school_id')
    radius = request.GET.get('radius', '500')
    year = request.GET.get('year', '')
    user_type = request.GET.get('user_type', 'propaganda')  # propaganda или organization
    
    if not school_id:
        return JsonResponse({'error': 'Укажите школу'}, status=400)
    
    try:
        school_id = int(school_id)
        radius = int(radius)
        year = int(year) if year else None
    except ValueError:
        return JsonResponse({'error': 'Неверные параметры'}, status=400)
    
    result = analyze_school(school_id, year=year, radius=radius)
    
    # Для обычного пользователя — упрощённый вывод
    if user_type == 'organization':
        return JsonResponse({
            'school_name': result.get('school_name', ''),
            'total_accidents': result.get('total_accidents', 0),
            'danger_level': result.get('danger_level', 'low'),
            'danger_level_name': result.get('danger_level_name', 'Низкий'),
            'recommendation': result.get('recommendation', {}),
            'severity_count': result.get('severity_count', {}),
            'categories': result.get('categories', []),  # Топ-3 категории
        })
    
    # Для ГИБДД — полный вывод
    return JsonResponse(result)

def ml_analysis_page(request):
    from datetime import datetime
    profile = request.user.profile
    is_propaganda = request.user.is_superuser or profile.user_type == 'propaganda'
    if not is_propaganda:
        return redirect('home')
    
    schools = Organization.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)
    
    years = list(range(2015, datetime.now().year + 1))
    years.reverse()
    
    return render(request, 'ml_analysis.html', {
        'all_schools': schools,
        'years': years,
        'is_propaganda': is_propaganda,
    })

@login_required
def district_report(request):
    """Отчёт: количество ДТП и мероприятий по районам"""
    if not (request.user.is_superuser or request.user.profile.user_type == 'propaganda'):
        messages.error(request, 'Нет доступа')
        return redirect('home')
    
    districts_list = [
        'Автовокзал', 'Академический', 'Березит', 'Ботанический', 'Верхнемакарово',
        'ВИЗ', 'Вокзальный', 'Втузгородок', 'Горный Щит', 'Елизавет', 'ЖБИ',
        'Завокзальный', 'Заречный', 'Изоплит', 'Исток', 'Калиновский', 'Карасьеозёрский',
        'Кольцово', 'Компрессорный', 'Лечебный', 'Малый Исток', 'Медный', 'Нижне-Исетский',
        'Новая Сортировка', 'Палкинский Торфяник', 'Палкино', 'Палникс', 'Парковый',
        'Пионерский', 'Полеводство', 'Птицефабрика', 'Рудный', 'Садовый', 'Северка',
        'Семь ключей', 'Сибирский', 'Синие Камни', 'Совхоз', 'Солнечный', 'Старая Сортировка',
        'Сулимовский', 'Технопарк', 'Уктус', 'УНЦ', 'Уралмаш', 'Химмаш', 'Центральный',
        'Чермет', 'Чусовское Озеро', 'Шабровский', 'Шарташ', 'Шарташский р-к',
        'Широкая Речка', 'Шувакиш', 'Эльмаш', 'Юго-Западный', 'Веер'
    ]
    
    year = request.GET.get('year', '')
    month = request.GET.get('month', '')
    selected_districts = request.GET.getlist('districts')
    
    try:
        year = int(year) if year else None
        month = int(month) if month else None
    except:
        year = None
        month = None
    
    if not year and not month and not selected_districts:
        available_years = list(range(2015, datetime.now().year + 1))[::-1]
        months = [(i, ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь'][i-1]) for i in range(1,13)]
        return render(request, 'district_report.html', {
            'district_data': None, 'total_dtp': 0, 'total_events': 0,
            'selected_year': '', 'selected_month': '', 'selected_districts': [],
            'available_years': available_years, 'months': months, 'districts_list': districts_list,
        })
    
    if not selected_districts:
        selected_districts = districts_list
    
    district_data = []
    total_dtp = 0
    total_events = 0
    
    for district in selected_districts:
        dtp_qs = Accident.objects.filter(district=district)
        events_qs = Event.objects.filter(district=district)
        
        if year:
            dtp_qs = dtp_qs.filter(date__year=year)
            events_qs = events_qs.filter(date__year=year)
        if month:
            dtp_qs = dtp_qs.filter(date__month=month)
            events_qs = events_qs.filter(date__month=month)
        
        dtp_count = dtp_qs.count()
        events_count = events_qs.count()
        
        if dtp_count > 0 or events_count > 0:
            # Детальный список ДТП
            dtp_list = []
            for a in dtp_qs.order_by('-date')[:50]:
                dtp_list.append({
                    'id': a.id,
                    'title': a.title,
                    'date': a.date.strftime('%d.%m.%Y'),
                    'severity': a.get_severity_display(),
                    'address': a.location_address or f"{a.latitude:.4f}, {a.longitude:.4f}" if a.latitude else '—',
                    'description': a.description or '',
                })
            
            # Детальный список мероприятий
            events_list = []
            for e in events_qs.order_by('-date')[:50]:
                events_list.append({
                    'id': e.id,
                    'title': e.title,
                    'date': e.date.strftime('%d.%m.%Y'),
                    'status': e.get_status_display(),
                    'type': e.get_event_type_display(),
                })
            
            district_data.append({
                'district': district,
                'dtp_total': dtp_count,
                'dtp_light': dtp_qs.filter(severity='light').count(),
                'dtp_heavy': dtp_qs.filter(severity='heavy').count(),
                'dtp_fatal': dtp_qs.filter(severity='fatal').count(),
                'events_total': events_count,
                'events_planned': events_qs.filter(status='planned').count(),
                'events_conducted': events_qs.filter(status='conducted').count(),
                'dtp_list': dtp_list,
                'events_list': events_list,
            })
            total_dtp += dtp_count
            total_events += events_count
    
    district_data.sort(key=lambda x: x['dtp_total'], reverse=True)
    
    # Экспорт в Excel (упрощённый)
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт по районам"
        period = f"{year or 'все годы'}{f' / месяц {month}' if month else ''}"
        ws.cell(row=1, column=1, value=f"Отчёт по районам за {period}").font = Font(bold=True, size=14)
        headers = ['Район', 'ДТП всего', 'Лёгких', 'Тяжёлых', 'Смертельных', 'Мероприятий', 'Запланировано', 'Проведено']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).font = Font(bold=True)
            ws.cell(row=3, column=col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
        for i, d in enumerate(district_data, 4):
            ws.cell(row=i, column=1, value=d['district'])
            ws.cell(row=i, column=2, value=d['dtp_total'])
            ws.cell(row=i, column=3, value=d['dtp_light'])
            ws.cell(row=i, column=4, value=d['dtp_heavy'])
            ws.cell(row=i, column=5, value=d['dtp_fatal'])
            ws.cell(row=i, column=6, value=d['events_total'])
            ws.cell(row=i, column=7, value=d['events_planned'])
            ws.cell(row=i, column=8, value=d['events_conducted'])
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="district_report.xlsx"'
        wb.save(response)
        return response
    
    available_years = list(range(2015, datetime.now().year + 1))[::-1]
    months = [(i, ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь'][i-1]) for i in range(1,13)]
    
    return render(request, 'district_report.html', {
        'district_data': district_data,
        'total_dtp': total_dtp,
        'total_events': total_events,
        'selected_year': year or '',
        'selected_month': month or '',
        'selected_districts': selected_districts,
        'available_years': available_years,
        'months': months,
        'districts_list': districts_list,
        'is_propaganda': True,
    })

def api_accidents_search(request):
    """Поиск ДТП по ключевым словам с фильтрацией по году и месяцам"""
    query = request.GET.get('q', '').strip()
    year = request.GET.get('year', '')
    months_str = request.GET.get('months', '')
    
    if not query or len(query) < 2:
        return JsonResponse([], safe=False)
    
    accidents = Accident.objects.filter(
        Q(title__icontains=query) |
        Q(description__icontains=query) |
        Q(location_address__icontains=query)
    ).filter(
        latitude__isnull=False, longitude__isnull=False
    ).exclude(latitude=0, longitude=0)
    
    if year and year.strip():
        accidents = accidents.filter(date__year=int(year))
    
    if months_str:
        months = [int(m) for m in months_str.split(',') if m.strip()]
        if months:
            accidents = accidents.filter(date__month__in=months)
    
    severity_map = {
        'light': 'Лёгкое', 'medium': 'Среднее',
        'heavy': 'Тяжёлое', 'fatal': 'Смертельное'
    }
    
    from .ml_utils import get_prevention_category
    
    result = []
    for a in accidents:
        category = get_prevention_category(f"{a.title} {a.description or ''} {a.location_address or ''}")
        result.append({
            'id': a.id,
            'title': a.title,
            'latitude': a.latitude,
            'longitude': a.longitude,
            'severity': a.severity,
            'severity_text': severity_map.get(a.severity, 'Неизвестно'),
            'date': a.date.strftime('%d.%m.%Y'),
            'address': a.location_address or '',
            'description': a.description or '',
            'category': category,
        })
    
    return JsonResponse(result, safe=False)

def api_accidents_tags(request):
    """Возвращает список уникальных категорий ДТП для автодополнения"""
    tags = set()
    for a in Accident.objects.exclude(description='').values_list('description', flat=True)[:5000]:
        import re
        for match in re.findall(r'Категория:\s*([^,\.]+)', str(a)):
            tags.add(match.strip())
    return JsonResponse(sorted(list(tags)), safe=False)